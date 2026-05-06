"""PMP Quiz Web Application"""
import os
import re
import json
import random
import bcrypt
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, abort
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from markupsafe import Markup, escape
from sqlalchemy import func, desc, and_, text

from config import Config
from models import db, User, Question, QuizSession, QuizAnswer, WrongAnswer, UserAnswerStat, QuestionGlobalStat, Bookmark, QuestionReport
from migrate import auto_migrate  # DB schema auto-sync (additive)

app = Flask(__name__)
app.config.from_object(Config)

# ══════════════════════════════════════════════════════
# 커스텀 도메인 redirect: *.up.railway.app → pmp.wayexam.com (301)
# ══════════════════════════════════════════════════════
PRIMARY_HOST = os.environ.get('PRIMARY_HOST', 'pmp.wayexam.com')


@app.before_request
def _redirect_to_primary_host():
    host = (request.host or '').lower()
    if (
        host
        and host != PRIMARY_HOST.lower()
        and not host.startswith('localhost')
        and not host.startswith('127.')
    ):
        target = 'https://' + PRIMARY_HOST + request.full_path.rstrip('?')
        return redirect(target, code=301)

# ══════════════════════════════════════════════════════
# Jinja filter: render markdown tables in question/explanation text
# Used by table-style questions (Q90001~90015) where the body
# contains a Markdown-style "| col | col |" table.
# Non-table text is HTML-escaped and \n is converted to <br>.
# ══════════════════════════════════════════════════════
_MD_TABLE_BLOCK = re.compile(
    r'(^[ \t]*\|[^\n]+\|[ \t]*\n'                   # header row
    r'[ \t]*\|[ \t\-:|]+\|[ \t]*\n'                  # separator row
    r'(?:[ \t]*\|[^\n]+\|[ \t]*(?:\n|$))+)',         # one or more body rows
    re.MULTILINE
)


def _md_table_to_html(block: str) -> str:
    lines = [ln.strip() for ln in block.strip().split('\n') if ln.strip()]
    if len(lines) < 2:
        return escape(block)

    def split_row(row: str):
        return [c.strip() for c in row.strip().strip('|').split('|')]

    header_cells = split_row(lines[0])
    body_rows = [split_row(r) for r in lines[2:]]

    out = ['<div class="table-wrapper" style="margin:10px 0;"><table class="md-table">']
    out.append('<thead><tr>')
    out.extend(f'<th>{escape(c)}</th>' for c in header_cells)
    out.append('</tr></thead><tbody>')
    for row in body_rows:
        out.append('<tr>')
        out.extend(f'<td>{escape(c)}</td>' for c in row)
        out.append('</tr>')
    out.append('</tbody></table></div>')
    return ''.join(out)


@app.template_filter('render_md_tables')
def render_md_tables(text_input):
    """Convert markdown tables in text to HTML; preserve newlines elsewhere."""
    if not text_input:
        return ''
    s = str(text_input)
    parts = []
    last = 0
    for m in _MD_TABLE_BLOCK.finditer(s):
        before = s[last:m.start()]
        if before:
            parts.append(str(escape(before)).replace('\n', '<br>'))
        parts.append(_md_table_to_html(m.group(1)))
        last = m.end()
    tail = s[last:]
    if tail:
        parts.append(str(escape(tail)).replace('\n', '<br>'))
    return Markup(''.join(parts))


db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to continue.'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated

# ══════════════════════════════════════════════════════
# Slicer / Filter Maps (NameError fix)
# ══════════════════════════════════════════════════════
HIERARCHY_PAIRS = [
    ('eco2021_domain', 'eco2021_task'),
    ('eco2026_domain', 'eco2026_task'),
    ('methodology',    'methodology_detail'),
]

FILTER_MAP = {
    'eco2021_domain':    Question.eco2021_domain,
    'eco2021_task':      Question.eco2021_task,
    'pmbok7_domain':     Question.pmbok7_domain,
    'pmbok7_principle':  Question.pmbok7_principle,
    'eco2026_domain':    Question.eco2026_domain,
    'eco2026_task':      Question.eco2026_task,
    'pmbok8_domain':     Question.pmbok8_domain,
    'pmbok8_principle':  Question.pmbok8_principle,
    'pmbok8_focus_area': Question.pmbok8_focus_area,
    'pmbok8_process':    Question.pmbok8_process,
    'pmbok8_new_topics': Question.pmbok8_new_topics,
    'methodology':       Question.methodology,
    'methodology_detail': Question.methodology_detail,
}


# ══════════════════════════════════════════════════════
# Lazy DB initialization
# ──────────────────────────────────────────────────────
# Heavy DB work (create_all, auto_migrate, admin seed, table-question seed,
# auto-load) is deferred to the first incoming request so that the gunicorn
# worker can bind to $PORT immediately. This prevents Railway's healthcheck
# from timing out (502) on cold start when DATABASE_URL points to a slow
# proxy connection.
# ══════════════════════════════════════════════════════
_DB_INITIALIZED = False


def _initialize_db_once():
    """Run startup DB tasks exactly once. Safe to call from request context."""
    global _DB_INITIALIZED
    if _DB_INITIALIZED:
        return
    try:
        db.create_all()
        auto_migrate(db)  # ALTER TABLE for columns added to models since last deploy
        # Create or promote admin users (idempotent)
        for email in Config.ADMIN_EMAILS:
            existing = User.query.filter_by(email=email).first()
            if not existing:
                admin = User(email=email, is_admin=True, is_premium=True)
                admin.set_validity(months=120)
                db.session.add(admin)
                print(f'[INIT] admin created: {email}')
            elif not existing.is_admin:
                existing.is_admin = True
                existing.is_premium = True
                if not existing.validity_end or existing.validity_end < datetime.utcnow():
                    existing.set_validity(months=120)
                print(f'[INIT] admin auto-promoted: {email}')
        db.session.commit()

        # Table Question 15 items seed (idempotent)
        try:
            from seed_table_questions import seed_table_questions
            seed_table_questions(db, Question)
        except Exception as _e:
            print(f'[INIT] Table Question seed failed: {_e}')

        # Auto-load questions if DB is empty
        if Question.query.count() == 0:
            filepath = 'data/PMP_Raw.xlsx'
            if os.path.exists(filepath):
                from load_data import load_questions
                count = load_questions(filepath)
                print(f"[STARTUP] Auto-loaded {count} questions from {filepath}")
            else:
                print(f"[STARTUP] No data file found at {filepath}")
        _DB_INITIALIZED = True
        print('[INIT] DB initialization complete.')
    except Exception as e:
        # Don't latch the flag on failure so a future request can retry.
        print(f'[INIT] DB initialization FAILED: {e}')
        raise


@app.before_request
def _ensure_db_initialized():
    """Lazy hook: initialize DB on first real request (not /healthz)."""
    if _DB_INITIALIZED:
        return
    # Skip init for the healthcheck endpoint so Railway can probe instantly.
    if request.path == '/healthz':
        return
    _initialize_db_once()


@app.route('/healthz')
def healthz():
    """Railway healthcheck — must respond instantly without touching DB."""
    return 'OK', 200

# ══════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        if not email or '@' not in email:
            flash('올바른 Email address를 입력해wk세요.', 'error')
            return render_template('login.html')
        if len(password) < 4:
            flash('Please enter your password (4+ chars).', 'error')
            return render_template('login.html', email=email)

        user = User.query.filter_by(email=email).first()
        if not user:
            flash('가입되지 않은 Email입니다. Sign up 후 이용해wk세요.', 'error')
            return redirect(url_for('signup', email=email))

        # Admin Email은 ADMIN_PASSWORD 로도 Log in 가능 (레거시 호환)
        is_admin_email = email in Config.ADMIN_EMAILS
        password_ok = False
        if user.password_hash:
            try:
                password_ok = bcrypt.checkpw(password.encode(), user.password_hash.encode())
            except Exception:
                password_ok = False
        if not password_ok and is_admin_email and password == Config.ADMIN_PASSWORD:
            # 첫 Admin Log in 시 Password 해시 Save
            user.password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
            db.session.commit()
            password_ok = True
        if not password_ok:
            flash('Email or password is incorrect.', 'error')
            return render_template('login.html', email=email)

        user.last_login = datetime.utcnow()
        db.session.commit()
        login_user(user, remember=True)
        return redirect(url_for('dashboard'))

    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        password2 = request.form.get('password2', '')
        if not email or '@' not in email:
            flash('올바른 Email address를 입력해wk세요.', 'error')
            return render_template('signup.html', email=email)
        if len(password) < 4:
            flash('Password는 최소 4+ chars이어야 합니다.', 'error')
            return render_template('signup.html', email=email)
        if password != password2:
            flash('Passwords do not match. 다시 OK해wk세요.', 'error')
            return render_template('signup.html', email=email)
        if User.query.filter_by(email=email).first():
            flash('already 가입된 Email입니다. Log in해wk세요.', 'error')
            return redirect(url_for('login', email=email))

        user = User(email=email)
        user.password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        user.is_premium = False
        is_admin_email = email in Config.ADMIN_EMAILS
        if is_admin_email:
            user.is_admin = True
            user.is_premium = True
            user.set_validity(months=120)
        user.last_login = datetime.utcnow()
        db.session.add(user)
        db.session.commit()
        # 신규가입 알림 메day (SMTP 설정 시 발송)
        try:
            _notify_signup(email)
        except Exception as e:
            app.logger.warning(f'[MAIL] signup notification failed: {e}')
        login_user(user, remember=True)
        flash('✅ 가입 Completed! 바로 Log in되었습니다.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('signup.html', email=request.args.get('email', ''))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

def _notify_signup(email):
    """신규가입 알림 메day 발송 (SMTP_* 환경변수 설정 시 동작)."""
    import os, smtplib
    from email.mime.text import MIMEText
    host = os.getenv('SMTP_HOST')
    user = os.getenv('SMTP_USER')
    pw = os.getenv('SMTP_PASS')
    if not (host and user and pw):
        now = datetime.utcnow()
        print(f'[MAIL][stub] PMP Quiz 아이디 {email} {now.month} mo{now.day}day 가입')
        return
    port = int(os.getenv('SMTP_PORT', '587'))
    to_addr = os.getenv('NOTIFY_EMAIL', 'songdoinfo@naver.com')
    now = datetime.utcnow()
    body = f'PMP Quiz 사이트 아이디 {email} {now.month} mo{now.day}day 가입'
    msg = MIMEText(body)
    msg['Subject'] = f'[PMP Quiz] 신규 가입: {email}'
    msg['From'] = os.getenv('SMTP_FROM', user)
    msg['To'] = to_addr
    with smtplib.SMTP(host, port) as s:
        s.starttls()
        s.login(user, pw)
        s.send_message(msg)

# ══════════════════════════════════════════════════════
# FREE VERSION (no login required)
# ══════════════════════════════════════════════════════

@app.route('/free')
def free_mode():
    already_used = session.get('free_used', False)
    return render_template('free_mode.html', already_used=already_used)

@app.route('/free/upgrade')
def free_upgrade():
    return render_template('free_upgrade.html')


# ══════════════════════════════════════════════════════
# LEMON SQUEEZY PAYMENT INTEGRATION
# ──────────────────────────────────────────────────────
# Flow:
#   1. User clicks "Buy Premium" -> /upgrade page (renders pricing).
#   2. User clicks a plan -> redirected to Lemon Squeezy hosted checkout.
#   3. User pays on Lemon Squeezy (cards / Apple Pay / etc.).
#   4. Lemon Squeezy fires order_created webhook to /webhook/lemonsqueezy.
#   5. Webhook handler verifies HMAC signature, looks up user by email,
#      sets premium validity, returns 200.
#   6. User is redirected to /payment/success with confirmation.
#
# Required env vars (set in Railway Variables when ready):
#   - LEMONSQUEEZY_STORE_SLUG       (e.g. 'pmp-quiz')
#   - LEMONSQUEEZY_WEBHOOK_SECRET   (used to verify signature)
#   - LEMONSQUEEZY_VARIANT_3MO      (variant ID for 3-month plan)
#   - LEMONSQUEEZY_VARIANT_6MO      (variant ID for 6-month plan, optional)
#   - LEMONSQUEEZY_VARIANT_12MO     (variant ID for 12-month plan, optional)
#
# Until env vars are set, /upgrade renders a "coming soon" notice and
# /webhook/lemonsqueezy returns 503. Site continues to work normally.
# ══════════════════════════════════════════════════════
import hmac
import hashlib

LEMONSQUEEZY_PLANS = [
    # (variant env var name, label, price USD, months of validity)
    ('LEMONSQUEEZY_VARIANT_3MO',  '3 Months Premium',  19, 3),
    ('LEMONSQUEEZY_VARIANT_6MO',  '6 Months Premium',  29, 6),
    ('LEMONSQUEEZY_VARIANT_12MO', '12 Months Premium', 49, 12),
]


def _lemonsqueezy_configured():
    """True only if minimum env vars are set."""
    return bool(
        os.environ.get('LEMONSQUEEZY_STORE_SLUG')
        and os.environ.get('LEMONSQUEEZY_WEBHOOK_SECRET')
        and os.environ.get('LEMONSQUEEZY_VARIANT_3MO')
    )


def _build_checkout_url(variant_id, email):
    """Build a Lemon Squeezy hosted checkout URL with prefilled email.
    Pattern: https://{store_slug}.lemonsqueezy.com/buy/{variant_uuid}?checkout[email]=...
    """
    store_slug = os.environ.get('LEMONSQUEEZY_STORE_SLUG', '')
    base = f'https://{store_slug}.lemonsqueezy.com/buy/{variant_id}'
    # Tag the email so we can match the user on webhook
    from urllib.parse import quote
    return f'{base}?checkout[email]={quote(email)}'


@app.route('/upgrade')
@login_required
def upgrade():
    """Render premium plans + checkout buttons."""
    plans = []
    for env_var, label, price, months in LEMONSQUEEZY_PLANS:
        variant_id = os.environ.get(env_var)
        if not variant_id:
            continue
        plans.append({
            'label': label,
            'price': price,
            'months': months,
            'checkout_url': _build_checkout_url(variant_id, current_user.email),
        })
    return render_template(
        'upgrade.html',
        plans=plans,
        configured=_lemonsqueezy_configured(),
    )


@app.route('/webhook/lemonsqueezy', methods=['POST'])
def webhook_lemonsqueezy():
    """Receive order_created event from Lemon Squeezy and grant premium.

    Lemon Squeezy webhook docs: signature is HMAC-SHA256 of raw body using
    the webhook secret, sent in X-Signature header (hex).
    """
    secret = os.environ.get('LEMONSQUEEZY_WEBHOOK_SECRET')
    if not secret:
        return 'webhook not configured', 503

    raw_body = request.get_data()
    received_sig = request.headers.get('X-Signature', '')
    expected_sig = hmac.new(
        secret.encode('utf-8'), raw_body, hashlib.sha256
    ).hexdigest()
    if not hmac.compare_digest(expected_sig, received_sig):
        print(f'[lemonsqueezy] BAD signature. expected={expected_sig[:12]}... got={received_sig[:12]}...')
        return 'bad signature', 401

    try:
        payload = json.loads(raw_body)
    except Exception as e:
        return f'bad json: {e}', 400

    event_name = payload.get('meta', {}).get('event_name', '')
    print(f'[lemonsqueezy] event={event_name}')

    # Only process successful orders (not subscription_payment_failed etc.)
    if event_name not in ('order_created', 'subscription_created', 'subscription_payment_success'):
        return 'event ignored', 200

    data = payload.get('data', {}).get('attributes', {})
    customer_email = (data.get('user_email') or data.get('customer_email') or '').strip().lower()
    if not customer_email:
        print('[lemonsqueezy] no customer email in payload')
        return 'no email', 400

    # Find which variant was purchased to determine months of validity
    variant_id = None
    if event_name == 'order_created':
        first_order_item = (data.get('first_order_item') or {})
        variant_id = str(first_order_item.get('variant_id') or '')
    if not variant_id:
        # subscription events
        variant_id = str(data.get('variant_id') or '')

    months = 3  # default
    for env_var, _label, _price, plan_months in LEMONSQUEEZY_PLANS:
        if str(os.environ.get(env_var, '')) == variant_id:
            months = plan_months
            break

    # Match user account (case-insensitive email)
    user = User.query.filter(func.lower(User.email) == customer_email).first()
    if not user:
        # Auto-create the account so the buyer can immediately log in.
        # (No password set; they'll need to use signup or password-reset.)
        user = User(email=customer_email, is_premium=True)
        user.set_validity(months=months)
        db.session.add(user)
        db.session.commit()
        print(f'[lemonsqueezy] auto-created premium user {customer_email} (+{months}mo)')
    else:
        user.is_premium = True
        user.extend_validity(months=months)
        db.session.commit()
        print(f'[lemonsqueezy] extended {customer_email} by {months} months '
              f'(new end: {user.validity_end})')

    return 'ok', 200


@app.route('/payment/success')
def payment_success():
    """User-facing redirect after successful checkout."""
    return render_template('payment_success.html')


@app.route('/payment/cancel')
def payment_cancel():
    flash('Payment was canceled. You can try again anytime.', 'info')
    return redirect(url_for('upgrade') if _lemonsqueezy_configured() else url_for('dashboard'))

@app.route('/free/start', methods=['POST'])
def free_start():
    # already Free trial을 Completed한 경우 Upgrade 페이지로 이동
    if session.get('free_used', False):
        return redirect(url_for('free_upgrade'))

    count = int(request.form.get('count', 10))
    count = min(count, Config.FREE_QUESTION_LIMIT)

    # 처음 N items 고정 Question (Random 아님, Question번호 순서대로)
    free_pool = Question.query.order_by(Question.no).limit(Config.FREE_QUESTION_LIMIT).all()
    selected = free_pool[:count]
    q_nos = [q.no for q in selected]

    session['free_questions'] = q_nos
    session['free_answers'] = {}
    session['free_current'] = 0
    return redirect(url_for('free_quiz'))

@app.route('/free/quiz')
def free_quiz():
    q_nos = session.get('free_questions', [])
    if not q_nos:
        return redirect(url_for('free_mode'))
    current = session.get('free_current', 0)
    if current >= len(q_nos):
        return redirect(url_for('free_result'))

    question = Question.query.filter_by(no=q_nos[current]).first()
    return render_template('free_quiz.html',
                         question=question,
                         current=current + 1,
                         total=len(q_nos),
                         answers=session.get('free_answers', {}))

@app.route('/free/answer', methods=['POST'])
def free_answer():
    q_no = request.form.get('question_no')
    selected = request.form.getlist('selected')
    answers = session.get('free_answers', {})
    answers[q_no] = ','.join(selected)
    session['free_answers'] = answers

    action = request.form.get('action', 'next')
    if action == 'prev':
        session['free_current'] = max(0, session.get('free_current', 0) - 1)
    elif action == 'next':
        session['free_current'] = session.get('free_current', 0) + 1
    elif action == 'goto':
        goto = int(request.form.get('goto_num', 0)) - 1
        session['free_current'] = max(0, min(goto, len(session.get('free_questions', [])) - 1))

    return redirect(url_for('free_quiz'))

@app.route('/free/grade', methods=['POST'])
def free_grade():
    q_nos = session.get('free_questions', [])
    answers = session.get('free_answers', {})

    # Save any last-minute answer
    q_no = request.form.get('question_no')
    selected = request.form.getlist('selected')
    if q_no and selected:
        answers[q_no] = ','.join(selected)
        session['free_answers'] = answers

    return redirect(url_for('free_result'))

@app.route('/free/result')
def free_result():
    q_nos = session.get('free_questions', [])
    user_answers = session.get('free_answers', {})
    if not q_nos:
        return redirect(url_for('free_mode'))

    questions = Question.query.filter(Question.no.in_(q_nos)).all()
    q_map = {q.no: q for q in questions}

    results = []
    correct_count = 0
    for no in q_nos:
        q = q_map.get(no)
        if not q:
            continue
        user_ans_raw = user_answers.get(str(no), '')
        user_ans = sorted([a.strip() for a in user_ans_raw.split(',') if a.strip()])
        correct_ans = sorted(q.get_answer_list())
        is_correct = user_ans == correct_ans
        if is_correct:
            correct_count += 1

        results.append({
            'question': q,
            'user_answer': user_ans,
            'correct_answer': correct_ans,
            'is_correct': is_correct,
            'answered': bool(user_ans_raw),
        })

    total = len(results)
    accuracy = (correct_count / total * 100) if total > 0 else 0

    # Free trial Completed 표시 (다시 Start 시 Upgrade 유도)
    session['free_used'] = True

    # Generate mock wrong-answer list and accuracy for free version
    return render_template('free_result.html',
                         results=results,
                         correct_count=correct_count,
                         total=total,
                         accuracy=accuracy)

# ══════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════

@app.route('/dashboard')
@login_required
def dashboard():
    # Recent sessions
    recent = QuizSession.query.filter_by(user_id=current_user.id, is_completed=True)\
        .order_by(desc(QuizSession.completed_at)).limit(5).all()

    # Overall stats
    total_sessions = QuizSession.query.filter_by(user_id=current_user.id, is_completed=True).count()
    avg_accuracy = db.session.query(func.avg(QuizSession.accuracy))\
        .filter_by(user_id=current_user.id, is_completed=True).scalar() or 0

    wrong_count = WrongAnswer.query.filter_by(user_id=current_user.id).count()
    total_questions = Question.query.count()

    # By category Select 섹션을 위한 categories context
    categories = get_category_options()

    return render_template('dashboard.html',
                         recent=recent,
                         total_sessions=total_sessions,
                         avg_accuracy=avg_accuracy,
                         wrong_count=wrong_count,
                         total_questions=total_questions,
                         categories=categories)

# ══════════════════════════════════════════════════════
# QUIZ ROUTES
# ══════════════════════════════════════════════════════

def _build_tree(parent_col, child_col):
    """Build {parent: [children]} dict, sorted."""
    rows = db.session.query(parent_col, child_col).filter(
        parent_col.isnot(None)
    ).distinct().order_by(parent_col, child_col).all()
    tree = {}
    for parent, child in rows:
        tree.setdefault(parent, [])
        if child and child not in tree[parent]:
            tree[parent].append(child)
    return tree

def get_category_options():
    """Get all available category options and hierarchical trees for filters"""
    return {
        'pmbok7': {
            'domain': sorted(set(q[0] for q in db.session.query(Question.pmbok7_domain).filter(Question.pmbok7_domain.isnot(None)).distinct().all())),
            'principle': sorted(set(q[0] for q in db.session.query(Question.pmbok7_principle).filter(Question.pmbok7_principle.isnot(None)).distinct().all())),
            'eco_domain': sorted(set(q[0] for q in db.session.query(Question.eco2021_domain).filter(Question.eco2021_domain.isnot(None)).distinct().all())),
            'eco_task': sorted(set(q[0] for q in db.session.query(Question.eco2021_task).filter(Question.eco2021_task.isnot(None)).distinct().all())),
        },
        'pmbok8': {
            'domain': sorted(set(q[0] for q in db.session.query(Question.pmbok8_domain).filter(Question.pmbok8_domain.isnot(None)).distinct().all())),
            'principle': sorted(set(q[0] for q in db.session.query(Question.pmbok8_principle).filter(Question.pmbok8_principle.isnot(None)).distinct().all())),
            'process': sorted(set(q[0] for q in db.session.query(Question.pmbok8_process).filter(Question.pmbok8_process.isnot(None)).distinct().all())),
            'focus_area': sorted(set(q[0] for q in db.session.query(Question.pmbok8_focus_area).filter(Question.pmbok8_focus_area.isnot(None)).distinct().all())),
            'new_topics': sorted(set(q[0] for q in db.session.query(Question.pmbok8_new_topics).filter(Question.pmbok8_new_topics.isnot(None)).distinct().all())),
            'eco_domain': sorted(set(q[0] for q in db.session.query(Question.eco2026_domain).filter(Question.eco2026_domain.isnot(None)).distinct().all())),
            'eco_task': sorted(set(q[0] for q in db.session.query(Question.eco2026_task).filter(Question.eco2026_task.isnot(None)).distinct().all())),
        },
        'methodology': sorted(set(q[0] for q in db.session.query(Question.methodology).filter(Question.methodology.isnot(None)).distinct().all())),
        # Hierarchical trees for cascading slicer UI
        'trees': {
            'eco2021': _build_tree(Question.eco2021_domain, Question.eco2021_task),
            'eco2026': _build_tree(Question.eco2026_domain, Question.eco2026_task),
            'methodology': _build_tree(Question.methodology, Question.methodology_detail),
        },
    }

@app.route('/quiz/start')
@login_required
def quiz_start():
    # Question풀이 탭은 dashboard로 통합됨
    return redirect(url_for('dashboard') + '#detail-select')


@app.route('/quiz/begin', methods=['POST'])
@login_required
def quiz_begin():
    mode = request.form.get('mode', 'random')
    filter_type = request.form.get('filter_type', '')
    filter_value = request.form.get('filter_value', '')
    count_raw = request.form.get('count', '10')
    count = int(count_raw) if count_raw.isdigit() else 10

    query = Question.query

    if mode == 'wrong_answers':
        wrong_nos = [w.question_no for w in WrongAnswer.query.filter_by(user_id=current_user.id).all()]
        if not wrong_nos:
            flash('Incorrect 목록이 비어있습니다.', 'info')
            return redirect(url_for('quiz_start'))
        query = query.filter(Question.no.in_(wrong_nos))

    elif mode == 'pmbok7_exam':
        count = 180
        query = query.filter(Question.pmbok7_domain.isnot(None))

    elif mode == 'pmbok8_exam':
        count = 185
        query = query.filter(Question.pmbok8_domain.isnot(None))

    elif mode == 'slicer':
        # 슬라이서 복수 Filter (계층 쌍은 OR, 나머지는 AND)
        filter_json_str = request.form.get('filter_json', '{}')
        try:
            filters = json.loads(filter_json_str)
        except Exception:
            filters = {}
        filter_type = 'slicer'
        filter_value = filter_json_str  # 세션 기록용

        from sqlalchemy import or_
        handled = set()
        for parent_key, child_key in HIERARCHY_PAIRS:
            parent_vals = filters.get(parent_key, [])
            child_vals  = filters.get(child_key, [])
            handled.update([parent_key, child_key])
            if parent_vals or child_vals:
                conds = []
                if parent_vals:
                    conds.append(FILTER_MAP[parent_key].in_(parent_vals))
                if child_vals:
                    conds.append(FILTER_MAP[child_key].in_(child_vals))
                query = query.filter(or_(*conds))

        for ft, values in filters.items():
            if ft in handled or not values:
                continue
            col = FILTER_MAP.get(ft)
            if col is not None:
                query = query.filter(col.in_(values))

    elif mode == 'category':
        if filter_type and filter_value:
            col = FILTER_MAP.get(filter_type)
            if col is not None:
                query = query.filter(col == filter_value)

    # Get questions
    all_questions = query.all()
    if not all_questions:
        flash('조건에 맞는 Question가 없습니다.', 'warning')
        return redirect(url_for('quiz_start'))

    random.shuffle(all_questions)
    selected = all_questions[:count]
    q_nos = [q.no for q in selected]

    # Create quiz session
    quiz_session = QuizSession(
        user_id=current_user.id,
        mode=mode,
        filter_type=filter_type,
        filter_value=filter_value,
        total_questions=len(q_nos),
    )
    db.session.add(quiz_session)
    db.session.commit()

    session['quiz_session_id'] = quiz_session.id
    session['quiz_questions'] = q_nos
    session['quiz_answers'] = {}
    session['quiz_current'] = 0

    return redirect(url_for('quiz_question'))

@app.route('/quiz/question')
@login_required
def quiz_question():
    q_nos = session.get('quiz_questions', [])
    if not q_nos:
        return redirect(url_for('quiz_start'))

    current = session.get('quiz_current', 0)
    if current >= len(q_nos):
        current = len(q_nos) - 1
        session['quiz_current'] = current

    question = Question.query.filter_by(no=q_nos[current]).first()
    saved_answer = session.get('quiz_answers', {}).get(str(q_nos[current]), '')

    is_bookmarked = bool(Bookmark.query.filter_by(
        user_id=current_user.id, question_no=q_nos[current]).first())

    return render_template('quiz_question.html',
                         question=question,
                         current=current + 1,
                         total=len(q_nos),
                         saved_answer=saved_answer.split(',') if saved_answer else [],
                         all_answers=session.get('quiz_answers', {}),
                         q_nos=q_nos,
                         is_bookmarked=is_bookmarked)

@app.route('/quiz/save_answer', methods=['POST'])
@login_required
def quiz_save_answer():
    q_no = request.form.get('question_no')
    selected = request.form.getlist('selected')
    answers = session.get('quiz_answers', {})
    if selected:
        answers[q_no] = ','.join(selected)
    session['quiz_answers'] = answers

    action = request.form.get('action', 'next')
    q_nos = session.get('quiz_questions', [])

    if action == 'prev':
        session['quiz_current'] = max(0, session.get('quiz_current', 0) - 1)
    elif action == 'next':
        session['quiz_current'] = min(len(q_nos) - 1, session.get('quiz_current', 0) + 1)
    elif action == 'goto':
        goto = int(request.form.get('goto_num', 1)) - 1
        session['quiz_current'] = max(0, min(goto, len(q_nos) - 1))
    elif action == 'grade':
        return redirect(url_for('quiz_grade'))

    return redirect(url_for('quiz_question'))

@app.route('/quiz/grade')
@login_required
def quiz_grade():
    quiz_session_id = session.get('quiz_session_id')
    q_nos = session.get('quiz_questions', [])
    user_answers = session.get('quiz_answers', {})

    if not quiz_session_id or not q_nos:
        return redirect(url_for('quiz_start'))

    quiz_sess = db.session.get(QuizSession, quiz_session_id)
    if not quiz_sess:
        return redirect(url_for('quiz_start'))

    questions = Question.query.filter(Question.no.in_(q_nos)).all()
    q_map = {q.no: q for q in questions}

    results = []
    correct_count = 0

    for no in q_nos:
        q = q_map.get(no)
        if not q:
            continue

        user_ans_raw = user_answers.get(str(no), '')
        user_ans = sorted([a.strip() for a in user_ans_raw.split(',') if a.strip()])
        correct_ans = sorted(q.get_answer_list())
        is_correct = user_ans == correct_ans
        if is_correct:
            correct_count += 1

        # Save quiz answer
        qa = QuizAnswer(
            session_id=quiz_session_id,
            question_no=no,
            user_answer=','.join(user_ans) if user_ans else '',
            correct_answer=q.answer,
            is_correct=is_correct,
        )
        db.session.add(qa)

        # Update wrong answers
        if is_correct:
            # Remove from wrong answers if exists
            WrongAnswer.query.filter_by(user_id=current_user.id, question_no=no).delete()
        else:
            wrong = WrongAnswer.query.filter_by(user_id=current_user.id, question_no=no).first()
            if wrong:
                wrong.wrong_count += 1
                wrong.last_wrong_at = datetime.utcnow()
            else:
                wrong = WrongAnswer(user_id=current_user.id, question_no=no)
                db.session.add(wrong)

        # Update user answer stats
        stat = UserAnswerStat.query.filter_by(user_id=current_user.id, question_no=no).first()
        if stat:
            stat.total_attempts += 1
            if is_correct:
                stat.correct_attempts += 1
            stat.last_attempted = datetime.utcnow()
        else:
            stat = UserAnswerStat(
                user_id=current_user.id,
                question_no=no,
                total_attempts=1,
                correct_attempts=1 if is_correct else 0,
            )
            db.session.add(stat)

        # Update global stats
        gstat = QuestionGlobalStat.query.filter_by(question_no=no).first()
        if gstat:
            gstat.total_attempts += 1
            if is_correct:
                gstat.correct_attempts += 1
            gstat.accuracy = (gstat.correct_attempts / gstat.total_attempts * 100) if gstat.total_attempts > 0 else 0
            gstat.last_updated = datetime.utcnow()
        else:
            gstat = QuestionGlobalStat(
                question_no=no,
                total_attempts=1,
                correct_attempts=1 if is_correct else 0,
                accuracy=100.0 if is_correct else 0.0,
            )
            db.session.add(gstat)

        results.append({
            'question': q,
            'user_answer': user_ans,
            'correct_answer': correct_ans,
            'is_correct': is_correct,
            'answered': bool(user_ans_raw),
        })

    # Update quiz session
    total = len(results)
    accuracy = (correct_count / total * 100) if total > 0 else 0
    quiz_sess.correct_count = correct_count
    quiz_sess.accuracy = accuracy
    quiz_sess.completed_at = datetime.utcnow()
    quiz_sess.is_completed = True
    db.session.commit()

    # Clear session quiz data
    for key in ['quiz_questions', 'quiz_answers', 'quiz_current', 'quiz_session_id']:
        session.pop(key, None)

    return render_template('quiz_result.html',
                         results=results,
                         correct_count=correct_count,
                         total=total,
                         accuracy=accuracy,
                         quiz_session=quiz_sess)

# ══════════════════════════════════════════════════════
# MY STATUS
# ══════════════════════════════════════════════════════

def _calc_streak(uid):
    """최근 풀이 day자 기반 Streakday 계산 (오늘  or  어제부터)"""
    from datetime import date, timedelta
    rows = db.session.query(func.distinct(func.date(QuizSession.completed_at)))\
        .filter_by(user_id=uid, is_completed=True).all()
    days = sorted({r[0] for r in rows if r[0]}, reverse=True)
    if not days:
        return 0
    today = date.today()
    cursor = today if days[0] == today else (today - timedelta(days=1))
    if days[0] != cursor:
        return 0
    streak = 0
    for d in days:
        if d == cursor:
            streak += 1
            cursor = cursor - timedelta(days=1)
        else:
            break
    return streak


def _calc_weak_domains(cat_stats, top_n=3):
    """전 카테고리에서 Accuracy 낮은 약점 Domain N items (시도 5회 이상). filter_json도 동봉."""
    import json as _json
    all_rows = []
    for ed in cat_stats.values():
        for key, rows in ed.items():
            for r in rows:
                if r.get('total', 0) >= 5:
                    all_rows.append({
                        'name': r['name'],
                        'accuracy': r['accuracy'],
                        'category': key,
                        'filter_json': _json.dumps({key: [r['name']]}, ensure_ascii=False),
                    })
    all_rows.sort(key=lambda x: x['accuracy'])
    return all_rows[:top_n]


def _sample_my_status_data():
    """Free user 미리View용 샘플 데이터 (재현 가능)"""
    from datetime import date, timedelta
    import random
    random.seed(42)
    today = date.today()
    daily = [{'date': str(today - timedelta(days=13 - i)),
              'avg_accuracy': round(random.uniform(58, 94), 1)} for i in range(14)]

    def cat(items, fkey):
        import json as _json
        return [{'name': n, 'correct': c, 'total': t, 'accuracy': round(c / t * 100, 1),
                 'filter_json': _json.dumps({fkey: [n]}, ensure_ascii=False)}
                for n, c, t in items]

    cat_stats = {
        'pmbok7': {
            'eco2021_domain': cat([('People', 35, 42), ('Process', 28, 38), ('Business Environment', 17, 25)], 'eco2021_domain'),
            'eco2021_task': cat([('Manage conflict', 8, 10), ('Engage stakeholders', 12, 15), ('Build a team', 9, 12)], 'eco2021_task'),
            'pmbok7_domain': cat([('Stakeholders', 14, 18), ('Team', 19, 22), ('Planning', 22, 30), ('Delivery', 17, 25)], 'pmbok7_domain'),
            'pmbok7_principle': cat([('Stewardship', 6, 8), ('Leadership', 10, 12), ('Tailoring', 5, 9)], 'pmbok7_principle'),
            'methodology': cat([('Agile', 18, 22), ('Waterfall', 12, 18), ('Hybrid', 8, 11)], 'methodology'),
        },
        'pmbok8': {
            'eco2026_domain': cat([('People', 32, 40), ('Process', 25, 36), ('Business Environment', 14, 22)], 'eco2026_domain'),
            'eco2026_task': cat([('Lead a team', 11, 14), ('Manage conflict', 9, 12)], 'eco2026_task'),
            'pmbok8_domain': cat([('Stakeholders', 13, 17), ('Team', 18, 21), ('Planning', 20, 28), ('Delivery', 15, 22)], 'pmbok8_domain'),
            'pmbok8_focus_area': cat([('AI/Automation', 5, 8), ('Sustainability', 6, 9), ('Diversity & Inclusion', 7, 10)], 'pmbok8_focus_area'),
            'pmbok8_process': cat([('Initiating', 8, 11), ('Planning', 18, 25), ('Executing', 14, 20), ('Monitoring & Controlling', 12, 18), ('Closing', 5, 7)], 'pmbok8_process'),
            'pmbok8_principle': cat([('Stewardship', 7, 10), ('Tailoring', 6, 9)], 'pmbok8_principle'),
            'methodology': cat([('Agile', 18, 22), ('Waterfall', 12, 18)], 'methodology'),
        },
    }
    return {
        'daily_stats': daily,
        'cat_stats': cat_stats,
        'wrong_count': 12,
        'total_attempted': 124,
        'total_correct': 91,
        'overall_accuracy': 73.4,
        'sessions_count': 9,
        'streak_days': 5,
        'weak_domains': [
            {'name': 'AI/Automation', 'accuracy': 62.5, 'category': 'pmbok8_focus_area', 'filter_json': '{"pmbok8_focus_area": ["AI/Automation"]}'},
            {'name': 'Business Environment', 'accuracy': 68.0, 'category': 'eco2021_domain', 'filter_json': '{"eco2021_domain": ["Business Environment"]}'},
            {'name': 'Sustainability', 'accuracy': 66.7, 'category': 'pmbok8_focus_area', 'filter_json': '{"pmbok8_focus_area": ["Sustainability"]}'},
        ],
    }


def _cat_stats(col, uid, filter_key=None):
    """min류 컬럼별 Accuracy 통계 helper. filter_key wk면 클릭용 filter_json도 동봉."""
    import json as _json
    rows = db.session.query(
        col,
        func.sum(UserAnswerStat.correct_attempts).label('correct'),
        func.sum(UserAnswerStat.total_attempts).label('total')
    ).join(UserAnswerStat, UserAnswerStat.question_no == Question.no)\
     .filter(UserAnswerStat.user_id == uid)\
     .filter(col.isnot(None))\
     .group_by(col)\
     .order_by(col).all()
    result = []
    for r in rows:
        total = int(r.total or 0)
        correct = int(r.correct or 0)
        name = r[0]
        result.append({
            'name': name,
            'total': total,
            'correct': correct,
            'accuracy': round(correct / total * 100, 1) if total > 0 else 0.0,
            'filter_json': _json.dumps({filter_key: [name]}, ensure_ascii=False) if filter_key else None,
        })
    return result

@app.route('/status')
@login_required
def my_status():
    uid = current_user.id

    validity_remaining = None
    if current_user.validity_end:
        delta = current_user.validity_end - datetime.utcnow()
        validity_remaining = max(0, delta.days)

    # Free user(미인증/Expired/Free등급) → 샘플 데이터로 미리View
    is_free_preview = (not current_user.is_admin) and (not current_user.is_premium or not current_user.is_valid())

    sessions = QuizSession.query.filter_by(user_id=uid, is_completed=True)\
        .order_by(desc(QuizSession.completed_at)).limit(50).all()

    # day별 추이
    daily_stats_raw = db.session.query(
        func.date(QuizSession.completed_at).label('date'),
        func.avg(QuizSession.accuracy).label('avg_accuracy'),
    ).filter_by(user_id=uid, is_completed=True)\
     .group_by(func.date(QuizSession.completed_at))\
     .order_by(func.date(QuizSession.completed_at)).all()
    daily_stats = [{'date': str(s.date), 'avg_accuracy': round(float(s.avg_accuracy), 1)} for s in daily_stats_raw]

    # All By category Accuracy
    cat_stats = {
        'pmbok7': {
            'eco2021_domain':   _cat_stats(Question.eco2021_domain,   uid, 'eco2021_domain'),
            'eco2021_task':     _cat_stats(Question.eco2021_task,     uid, 'eco2021_task'),
            'pmbok7_domain':    _cat_stats(Question.pmbok7_domain,    uid, 'pmbok7_domain'),
            'pmbok7_principle': _cat_stats(Question.pmbok7_principle, uid, 'pmbok7_principle'),
            'methodology':      _cat_stats(Question.methodology,      uid, 'methodology'),
        },
        'pmbok8': {
            'eco2026_domain':    _cat_stats(Question.eco2026_domain,    uid, 'eco2026_domain'),
            'eco2026_task':      _cat_stats(Question.eco2026_task,      uid, 'eco2026_task'),
            'pmbok8_domain':     _cat_stats(Question.pmbok8_domain,     uid, 'pmbok8_domain'),
            'pmbok8_focus_area': _cat_stats(Question.pmbok8_focus_area, uid, 'pmbok8_focus_area'),
            'pmbok8_process':    _cat_stats(Question.pmbok8_process,    uid, 'pmbok8_process'),
            'pmbok8_principle':  _cat_stats(Question.pmbok8_principle,  uid, 'pmbok8_principle'),
            'methodology':       _cat_stats(Question.methodology,       uid, 'methodology'),
        },
    }

    wrong_count = WrongAnswer.query.filter_by(user_id=uid).count()
    total_attempted = db.session.query(func.sum(UserAnswerStat.total_attempts))\
        .filter_by(user_id=uid).scalar() or 0
    total_correct = db.session.query(func.sum(UserAnswerStat.correct_attempts))\
        .filter_by(user_id=uid).scalar() or 0
    overall_accuracy = round(total_correct / total_attempted * 100, 1) if total_attempted > 0 else 0.0

    # streak / 약점 Domain
    streak_days = _calc_streak(uid)
    weak_domains = _calc_weak_domains(cat_stats, top_n=3)
    sessions_count = len(sessions)

    # Free user라면 샘플로 치환 (실 데이터가 빈약해도 와우 효과)
    if is_free_preview and (total_attempted == 0 or sessions_count == 0):
        sample = _sample_my_status_data()
        daily_stats = sample['daily_stats']
        cat_stats = sample['cat_stats']
        wrong_count = sample['wrong_count']
        total_attempted = sample['total_attempted']
        total_correct = sample['total_correct']
        overall_accuracy = sample['overall_accuracy']
        sessions_count = sample['sessions_count']
        streak_days = sample['streak_days']
        weak_domains = sample['weak_domains']
        sample_mode = True
    else:
        sample_mode = False

    return render_template('my_status.html',
                           validity_remaining=validity_remaining,
                           sessions=sessions,
                           daily_stats=daily_stats,
                           cat_stats=cat_stats,
                           wrong_count=wrong_count,
                           total_attempted=total_attempted,
                           total_correct=total_correct,
                           overall_accuracy=overall_accuracy,
                           streak_days=streak_days,
                           weak_domains=weak_domains,
                           sessions_count=sessions_count,
                           sample_mode=sample_mode,
                           is_free_preview=is_free_preview)

# ══════════════════════════════════════════════════════
# API ENDPOINTS
# ══════════════════════════════════════════════════════

@app.route('/api/framework_stats')
@login_required
def api_framework_stats():
    """Return user accuracy + question counts per classification value, grouped by dimension.

    Used by the dashboard's framework grid to color-code each classification box
    by the user's strength in that area and display question availability.
    """
    uid = current_user.id

    def values_for(col):
        all_values = sorted(
            v[0] for v in db.session.query(col).filter(col.isnot(None)).distinct().all()
        )
        stats_rows = db.session.query(
            col,
            func.sum(UserAnswerStat.correct_attempts).label('correct'),
            func.sum(UserAnswerStat.total_attempts).label('total')
        ).join(UserAnswerStat, UserAnswerStat.question_no == Question.no)\
         .filter(UserAnswerStat.user_id == uid)\
         .filter(col.isnot(None))\
         .group_by(col).all()
        stats_map = {r[0]: (int(r.total or 0), int(r.correct or 0)) for r in stats_rows}
        cap_rows = db.session.query(col, func.count(Question.id))\
            .filter(col.isnot(None)).group_by(col).all()
        cap_map = {r[0]: int(r[1]) for r in cap_rows}
        out = []
        for v in all_values:
            total, correct = stats_map.get(v, (0, 0))
            out.append({
                'name': v,
                'total_attempts': total,
                'correct_attempts': correct,
                'accuracy': round(correct / total * 100, 1) if total > 0 else None,
                'question_count': cap_map.get(v, 0),
            })
        return out

    return jsonify({
        'pmbok7_domain': values_for(Question.pmbok7_domain),
        'pmbok7_principle': values_for(Question.pmbok7_principle),
        'pmbok8_domain': values_for(Question.pmbok8_domain),
        'pmbok8_principle': values_for(Question.pmbok8_principle),
        'pmbok8_focus_area': values_for(Question.pmbok8_focus_area),
        'pmbok8_process': values_for(Question.pmbok8_process),
        'eco2021_domain': values_for(Question.eco2021_domain),
        'eco2021_task': values_for(Question.eco2021_task),
        'eco2026_domain': values_for(Question.eco2026_domain),
        'eco2026_task': values_for(Question.eco2026_task),
        'methodology': values_for(Question.methodology),
    })


@app.route('/api/subcategories')
@login_required
def api_subcategories():
    """Get subcategories based on filter type"""
    filter_type = request.args.get('filter_type', '')
    filter_map = {
        'pmbok7_domain': Question.pmbok7_domain,
        'pmbok7_principle': Question.pmbok7_principle,
        'eco2021_domain': Question.eco2021_domain,
        'eco2021_task': Question.eco2021_task,
        'pmbok8_domain': Question.pmbok8_domain,
        'pmbok8_principle': Question.pmbok8_principle,
        'pmbok8_process': Question.pmbok8_process,
        'pmbok8_focus_area': Question.pmbok8_focus_area,
        'pmbok8_new_topics': Question.pmbok8_new_topics,
        'eco2026_domain': Question.eco2026_domain,
        'eco2026_task': Question.eco2026_task,
        'methodology': Question.methodology,
    }
    col = filter_map.get(filter_type)
    if col is None:
        return jsonify([])

    values = sorted(set(q[0] for q in db.session.query(col).filter(col.isnot(None)).distinct().all()))
    counts = {}
    for val in values:
        counts[val] = Question.query.filter(col == val).count()
    return jsonify([{'value': v, 'count': counts.get(v, 0)} for v in values])

@app.route('/api/filter_count', methods=['POST'])
@login_required
def api_filter_count():
    """슬라이서 복수 Filter → this question 수 반환 (계층 쌍은 OR)"""
    try:
        filters = request.get_json(force=True) or {}
    except Exception:
        return jsonify({'count': 0})
    from sqlalchemy import or_
    query = Question.query
    handled = set()
    for parent_key, child_key in HIERARCHY_PAIRS:
        parent_vals = filters.get(parent_key, [])
        child_vals  = filters.get(child_key, [])
        handled.update([parent_key, child_key])
        if parent_vals or child_vals:
            conds = []
            if parent_vals:
                conds.append(FILTER_MAP[parent_key].in_(parent_vals))
            if child_vals:
                conds.append(FILTER_MAP[child_key].in_(child_vals))
            query = query.filter(or_(*conds))
    for ft, values in filters.items():
        if ft in handled or not values:
            continue
        col = FILTER_MAP.get(ft)
        if col is not None:
            query = query.filter(col.in_(values))
    return jsonify({'count': query.count()})

@app.route('/api/daily_trend')
@login_required
def api_daily_trend():
    stats = db.session.query(
        func.date(QuizSession.completed_at).label('date'),
        func.avg(QuizSession.accuracy).label('avg_accuracy'),
    ).filter_by(user_id=current_user.id, is_completed=True)\
     .group_by(func.date(QuizSession.completed_at))\
     .order_by(func.date(QuizSession.completed_at)).all()

    return jsonify([{
        'date': str(s.date),
        'accuracy': round(s.avg_accuracy, 1),
    } for s in stats])

# ══════════════════════════════════════════════════════
# ADMIN ROUTES
# ══════════════════════════════════════════════════════

@app.route('/admin')
@admin_required
def admin_panel():
    sort = request.args.get('sort', 'last_login')
    order = request.args.get('order', 'desc')
    filter_grade = request.args.get('grade', 'all')  # all / premium / free

    q = User.query
    if filter_grade == 'premium':
        q = q.filter_by(is_premium=True)
    elif filter_grade == 'free':
        q = q.filter_by(is_premium=False)

    sort_map = {
        'email':        (User.email,        'asc'),
        'grade':        (User.is_premium,   'desc'),
        'validity':     (User.validity_end, 'desc'),
        'last_login':   (User.last_login,   'desc'),
    }
    col, default_order = sort_map.get(sort, (User.last_login, 'desc'))
    actual_order = order if order in ('asc', 'desc') else default_order
    users = q.order_by(col.asc() if actual_order == 'asc' else col.desc()).all()

    total_questions = Question.query.count()
    pending_report_count = QuestionReport.query.filter_by(status='pending').count()
    return render_template('admin.html',
                           users=users,
                           total_questions=total_questions,
                           sort=sort, order=order, filter_grade=filter_grade,
                           pending_report_count=pending_report_count)

@app.route('/admin/import_translations', methods=['GET', 'POST'])
@admin_required
def admin_import_translations():
    """Merge zh/es/ja translations from data/PMP_Raw_translated.xlsx
    (sheet 'PMP_Translated') into Question rows by `no`.
    Idempotent — running again just overwrites with the latest values.

    Returns plain-text response so any error shows the full traceback in the
    browser instead of a generic 500 page (and avoids any redirect/template
    surprises while debugging)."""
    import traceback
    filepath = 'data/PMP_Raw_translated.xlsx'
    abs_path = os.path.abspath(filepath)
    cwd = os.getcwd()
    try:
        if not os.path.exists(filepath):
            return (f"File not found.\ncwd={cwd}\nlooking for: {abs_path}\n"
                    f"data/ contents: {os.listdir('data') if os.path.isdir('data') else 'no data/ dir'}\n"), 200, {'Content-Type': 'text/plain; charset=utf-8'}

        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        if 'PMP_Translated' not in wb.sheetnames:
            return (f"Sheet 'PMP_Translated' not found.\nSheets: {wb.sheetnames}\n"), 200, {'Content-Type': 'text/plain; charset=utf-8'}
        ws = wb['PMP_Translated']

        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = {name: i for i, name in enumerate(header) if name}

        field_map = {
            'Question_ZH': 'question_zh', 'OptA_ZH': 'opt_a_zh', 'OptB_ZH': 'opt_b_zh',
            'OptC_ZH': 'opt_c_zh', 'OptD_ZH': 'opt_d_zh', 'OptE_ZH': 'opt_e_zh',
            'Explanation_ZH': 'explanation_zh',
            'Question_ES': 'question_es', 'OptA_ES': 'opt_a_es', 'OptB_ES': 'opt_b_es',
            'OptC_ES': 'opt_c_es', 'OptD_ES': 'opt_d_es', 'OptE_ES': 'opt_e_es',
            'Explanation_ES': 'explanation_es',
            'Question_JA': 'question_ja', 'OptA_JA': 'opt_a_ja', 'OptB_JA': 'opt_b_ja',
            'OptC_JA': 'opt_c_ja', 'OptD_JA': 'opt_d_ja', 'OptE_JA': 'opt_e_ja',
            'Explanation_JA': 'explanation_ja',
        }
        no_col = col_idx.get('No')
        if no_col is None:
            return (f"'No' column missing.\nHeader: {header}\n"), 200, {'Content-Type': 'text/plain; charset=utf-8'}

        updated = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            qno = row[no_col]
            if not qno:
                continue
            q = Question.query.filter_by(no=qno).first()
            if not q:
                skipped += 1
                continue
            for xlsx_col, field in field_map.items():
                ci = col_idx.get(xlsx_col)
                if ci is None:
                    continue
                val = row[ci]
                if val is not None and val != '':
                    setattr(q, field, str(val))
            updated += 1
        db.session.commit()
        return (f"OK. updated={updated}, skipped={skipped} (unmatched no).\n"
                f"Header keys mapped: {sorted(set(field_map) & set(col_idx))}\n"), 200, {'Content-Type': 'text/plain; charset=utf-8'}
    except Exception as e:
        db.session.rollback()
        tb = traceback.format_exc()
        return (f"FAILED: {type(e).__name__}: {e}\n\n{tb}\n"), 200, {'Content-Type': 'text/plain; charset=utf-8'}


@app.route('/admin/reload_questions', methods=['GET', 'POST'])
@admin_required
def admin_reload_questions():
    """Force-reload questions from data/PMP_Raw.xlsx.
    Idempotent: load_data.py upserts by question.no, so existing rows
    are updated and new rows inserted. Safe to call multiple times.
    Useful for the EN site whose initial DB only has the 15 seed table
    questions; this populates the full 2,234-question dataset from the
    xlsx that was committed to data/."""
    filepath = 'data/PMP_Raw.xlsx'
    if not os.path.exists(filepath):
        flash(f'File not found: {filepath}', 'error')
        return redirect(url_for('admin_panel'))
    try:
        from load_data import load_questions
        count = load_questions(filepath)
        flash(f'Reloaded {count} questions from {filepath}.', 'success')
    except Exception as e:
        flash(f'Reload failed: {e}', 'error')
    return redirect(url_for('admin_panel'))


@app.route('/admin/add_user', methods=['POST'])
@admin_required
def admin_add_user():
    email = request.form.get('email', '').strip().lower()
    months = int(request.form.get('months', 3))
    is_premium = request.form.get('is_premium') == '1'

    if not email or '@' not in email:
        flash('올바른 Email address를 입력해wk세요.', 'error')
        return redirect(url_for('admin_panel'))

    if User.query.filter_by(email=email).first():
        flash(f'{email} 은 already 등록된 사용자입니다.', 'warning')
        return redirect(url_for('admin_panel'))

    user = User(email=email, is_premium=is_premium)
    user.set_validity(months=months)
    if email in Config.ADMIN_EMAILS:
        user.is_admin = True
        user.is_premium = True
        user.set_validity(months=120)
    db.session.add(user)
    db.session.commit()
    flash(f'{email} 사용자가 Added. (Valid until {months} mo)', 'success')
    return redirect(url_for('admin_panel'))


@app.route('/admin/user/<int:user_id>', methods=['POST'])
@admin_required
def admin_update_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash('사용자를 찾을 수 없습니다.', 'error')
        return redirect(url_for('admin_panel'))

    action = request.form.get('action')

    if action == 'set_validity':
        months = int(request.form.get('months', 3))
        user.set_validity(months)
        user.is_premium = True
        flash(f'{user.email}의 Valid until을 오늘부터 {months} mo로 설정했습니다.', 'success')

    elif action == 'extend_validity':
        months = int(request.form.get('months', 3))
        user.extend_validity(months)
        user.is_premium = True
        end_date = user.validity_end.strftime('%Y-%m-%d') if user.validity_end else '?'
        flash(f'{user.email}의 Valid until을 {months} mo 연장했습니다. (Expiry: {end_date})', 'success')

    elif action == 'toggle_premium':
        user.is_premium = not user.is_premium
        if user.is_premium and not user.validity_end:
            user.set_validity(months=Config.DEFAULT_VALIDITY_MONTHS)
        flash(f'{user.email}의 Premium 상태를 변경했습니다.', 'success')

    elif action == 'toggle_admin':
        if user.id != current_user.id:
            user.is_admin = not user.is_admin
            flash(f'{user.email}의 Admin 상태를 변경했습니다.', 'success')

    elif action == 'delete':
        if user.id != current_user.id:
            db.session.delete(user)
            flash(f'{user.email} 계정을 Delete했습니다.', 'success')

    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/questions')
@admin_required
def admin_questions():
    """Question 검색 & 목록"""
    search = request.args.get('q', '').strip()
    page = int(request.args.get('page', 1))
    per_page = 20

    query = Question.query
    if search:
        if search.isdigit():
            query = query.filter(Question.no == int(search))
        else:
            like = f'%{search}%'
            query = query.filter(
                Question.question.ilike(like) |
                Question.question_kr.ilike(like) |
                Question.explanation.ilike(like)
            )

    total = query.count()
    questions = query.order_by(Question.no).offset((page - 1) * per_page).limit(per_page).all()
    total_pages = (total + per_page - 1) // per_page

    return render_template('admin_questions.html',
                           questions=questions, search=search,
                           page=page, total_pages=total_pages, total=total)


@app.route('/admin/questions/<int:q_no>/edit', methods=['GET', 'POST'])
@admin_required
def admin_question_edit(q_no):
    """Question 직접 Edit"""
    q = Question.query.filter_by(no=q_no).first_or_404()

    if request.method == 'POST':
        # 영문 내용
        q.question    = request.form.get('question', '').strip()
        q.opt_a       = request.form.get('opt_a', '').strip() or None
        q.opt_b       = request.form.get('opt_b', '').strip() or None
        q.opt_c       = request.form.get('opt_c', '').strip() or None
        q.opt_d       = request.form.get('opt_d', '').strip() or None
        q.opt_e       = request.form.get('opt_e', '').strip() or None
        q.answer      = request.form.get('answer', '').strip().upper()
        q.explanation = request.form.get('explanation', '').strip() or None
        # 한국어 번역
        q.question_kr    = request.form.get('question_kr', '').strip() or None
        q.opt_a_kr       = request.form.get('opt_a_kr', '').strip() or None
        q.opt_b_kr       = request.form.get('opt_b_kr', '').strip() or None
        q.opt_c_kr       = request.form.get('opt_c_kr', '').strip() or None
        q.opt_d_kr       = request.form.get('opt_d_kr', '').strip() or None
        q.opt_e_kr       = request.form.get('opt_e_kr', '').strip() or None
        q.explanation_kr = request.form.get('explanation_kr', '').strip() or None
        # min류
        q.pmbok7_domain    = request.form.get('pmbok7_domain', '').strip() or None
        q.pmbok7_principle = request.form.get('pmbok7_principle', '').strip() or None
        q.pmbok8_domain    = request.form.get('pmbok8_domain', '').strip() or None
        q.pmbok8_principle = request.form.get('pmbok8_principle', '').strip() or None
        q.pmbok8_process   = request.form.get('pmbok8_process', '').strip() or None
        q.pmbok8_focus_area  = request.form.get('pmbok8_focus_area', '').strip() or None
        q.eco2021_domain   = request.form.get('eco2021_domain', '').strip() or None
        q.eco2021_task     = request.form.get('eco2021_task', '').strip() or None
        q.eco2026_domain   = request.form.get('eco2026_domain', '').strip() or None
        q.eco2026_task     = request.form.get('eco2026_task', '').strip() or None
        q.methodology      = request.form.get('methodology', '').strip() or None

        db.session.commit()
        flash(f'Question {q_no}번이 수정되었습니다.', 'success')

        next_page = request.form.get('next', '')
        return redirect(next_page if next_page else url_for('admin_questions'))

    # GET: min류 드롭다운 옵션용
    categories = get_category_options()
    back = request.args.get('back', url_for('admin_questions'))
    return render_template('admin_question_edit.html', q=q, categories=categories, back=back)


@app.route('/admin/question_stats')
@admin_required
def admin_question_stats():
    """Admin용 Per-question accuracy"""
    stats = db.session.query(
        QuestionGlobalStat,
        Question
    ).join(Question, Question.no == QuestionGlobalStat.question_no)\
     .order_by(QuestionGlobalStat.accuracy).all()

    return render_template('admin_question_stats.html', stats=stats)

@app.route('/admin/toggle_payment', methods=['POST'])
@admin_required
def admin_toggle_payment():
    app.config['PAYMENT_ENABLED'] = not app.config.get('PAYMENT_ENABLED', False)
    status = '활성화' if app.config['PAYMENT_ENABLED'] else '비활성화'
    flash(f'Payment이 {status}되었습니다.', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/load_data', methods=['POST'])
@admin_required
def admin_load_data():
    """Trigger data reload from Excel"""
    filepath = 'data/PMP_Raw.xlsx'
    if os.path.exists(filepath):
        from load_data import load_questions
        count = load_questions(filepath)
        flash(f'{count} items의 Question를 로드했습니다.', 'success')
    else:
        flash('Excel 파day을 찾을 수 없습니다.', 'error')
    return redirect(url_for('admin_panel'))

# ══════════════════════════════════════════════════════
# CONTEXT PROCESSORS
# ══════════════════════════════════════════════════════

@app.context_processor
def inject_config():
    return {
        'payment_enabled': app.config.get('PAYMENT_ENABLED', False),
        'contact_email': 'songdoinfo@naver.com',
    }

# ══════════════════════════════════════════════════════
# BOOKMARK ROUTES
# ══════════════════════════════════════════════════════

@app.route('/bookmarks')
@login_required
def bookmarks():
    items = (Bookmark.query
             .filter_by(user_id=current_user.id)
             .order_by(Bookmark.created_at.desc())
             .all())
    return render_template('bookmarks.html', bookmarks=items)

@app.route('/api/bookmark/toggle', methods=['POST'])
@login_required
def api_bookmark_toggle():
    data = request.get_json()
    q_no = data.get('question_no')
    if not q_no:
        return jsonify({'error': 'missing question_no'}), 400

    existing = Bookmark.query.filter_by(user_id=current_user.id, question_no=q_no).first()
    if existing:
        db.session.delete(existing)
        db.session.commit()
        count = Bookmark.query.filter_by(user_id=current_user.id).count()
        return jsonify({'bookmarked': False, 'count': count})
    else:
        bm = Bookmark(user_id=current_user.id, question_no=q_no)
        db.session.add(bm)
        db.session.commit()
        count = Bookmark.query.filter_by(user_id=current_user.id).count()
        return jsonify({'bookmarked': True, 'count': count})

# ══════════════════════════════════════════════════════
# REPORT ROUTES
# ══════════════════════════════════════════════════════

@app.route('/api/report', methods=['POST'])
@login_required
def api_report():
    data = request.get_json()
    q_no   = data.get('question_no')
    reason = data.get('reason', 'other')
    detail = data.get('detail', '')
    if not q_no:
        return jsonify({'error': 'missing question_no'}), 400

    # 동day 유저·동day Question 중복 pending Report 방지
    existing = QuestionReport.query.filter_by(
        user_id=current_user.id, question_no=q_no, status='pending').first()
    if existing:
        return jsonify({'ok': False, 'msg': 'already Report한 Question입니다.'})

    rpt = QuestionReport(user_id=current_user.id, question_no=q_no,
                         reason=reason, detail=detail)
    db.session.add(rpt)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/admin/reports')
@login_required
@admin_required
def admin_reports():
    status_filter = request.args.get('status', 'pending')
    reports = (QuestionReport.query
               .filter_by(status=status_filter)
               .order_by(QuestionReport.created_at.desc())
               .all())
    pending_count = QuestionReport.query.filter_by(status='pending').count()
    return render_template('admin_reports.html',
                           reports=reports,
                           status_filter=status_filter,
                           pending_count=pending_count)

@app.route('/admin/reports/<int:report_id>/resolve', methods=['POST'])
@login_required
@admin_required
def admin_report_resolve(report_id):
    rpt = QuestionReport.query.get_or_404(report_id)
    rpt.status = 'resolved'
    rpt.resolved_at = datetime.utcnow()
    db.session.commit()
    flash('Report가 처리됐습니다.', 'success')
    return redirect(url_for('admin_reports'))

@app.route('/admin/reports/<int:report_id>/dismiss', methods=['POST'])
@login_required
@admin_required
def admin_report_dismiss(report_id):
    rpt = QuestionReport.query.get_or_404(report_id)
    rpt.status = 'dismissed'
    rpt.resolved_at = datetime.utcnow()
    db.session.commit()
    flash('Report가 무시됐습니다.', 'info')
    return redirect(url_for('admin_reports'))

# ══════════════════════════════════════════════════════
# ERROR HANDLERS
# ══════════════════════════════════════════════════════

@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403, message='접근 권한이 없습니다.'), 403

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, message='페이지를 찾을 수 없습니다.'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('error.html', code=500, message='서버 An error occurred.'), 500

# ══════════════════════════════════════════════════════
# SEO ROUTES
# ══════════════════════════════════════════════════════

@app.route('/robots.txt')
def robots_txt():
    """Serve robots.txt for SEO"""
    try:
        with open(os.path.join(app.static_folder, 'robots.txt'), 'r') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/plain'}
    except FileNotFoundError:
        abort(404)

@app.route('/sitemap.xml')
def sitemap_xml():
    """Serve sitemap.xml for SEO"""
    try:
        with open(os.path.join(app.static_folder, 'sitemap.xml'), 'r') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'application/xml'}
    except FileNotFoundError:
        abort(404)

@app.route('/llms.txt')
def llms_txt():
    """Serve llms.txt for AI indexing"""
    try:
        with open(os.path.join(app.static_folder, 'llms.txt'), 'r') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/plain'}
    except FileNotFoundError:
        abort(404)

@app.route('/ads.txt')
def ads_txt():
    """Serve ads.txt for Google AdSense publisher verification.
    Format: 'google.com, pub-XXXXXXXXXXXXXXXX, DIRECT, f08c47fec0942fa0'
    Generated dynamically from ADSENSE_PUBLISHER_ID env var.
    """
    publisher = app.config.get('ADSENSE_PUBLISHER_ID', '')
    if not publisher:
        abort(404)
    pub_id = publisher.replace('ca-pub-', 'pub-') if publisher.startswith('ca-pub-') else publisher
    body = f'google.com, {pub_id}, DIRECT, f08c47fec0942fa0\n'
    return body, 200, {'Content-Type': 'text/plain'}

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)


# i18n: language resolver (added by transform.py)
# 'ko' is the secondary translation (Korean) — DB columns use legacy `_kr` suffix.
SUPPORTED_LANGS = ('en', 'ko', 'zh', 'es', 'ja')

@app.before_request
def _resolve_lang():
    from flask import g, request
    lang = request.cookies.get('lang', 'en')
    if lang not in SUPPORTED_LANGS:
        lang = 'en'
    g.lang = lang

@app.context_processor
def _inject_lang():
    from flask import g
    return {'current_lang': getattr(g, 'lang', 'en'),
            'supported_langs': SUPPORTED_LANGS}

@app.route('/lang/<lang>')
def set_lang(lang):
    from flask import redirect, request, make_response
    if lang not in SUPPORTED_LANGS:
        lang = 'en'
    resp = make_response(redirect(request.referrer or '/'))
    resp.set_cookie('lang', lang, max_age=60*60*24*365, samesite='Lax')
    return resp

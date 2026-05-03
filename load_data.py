"""Load PMP questions from Excel file(s) into database.

Supports an optional second xlsx (translated.xlsx) that adds Chinese,
Spanish, Japanese columns keyed on question No. Both files merged on No.
"""
import sys
import os
from openpyxl import load_workbook
from app import app
from models import db, Question

# Column mapping from main Excel to model fields
COLUMN_MAP = {
    'No': 'no',
    'Question. 번역, 코드,Web으로 만들기': 'question',
    'A': 'opt_a',
    'B': 'opt_b',
    'C': 'opt_c',
    'D': 'opt_d',
    'E': 'opt_e',
    'Explanation(Changed)': 'explanation',
    '2021 ECO Domain': 'eco2021_domain',
    '2021 ECO Task': 'eco2021_task',
    'PMBOK7 Performance Domain': 'pmbok7_domain',
    'PMBOK7 Principle': 'pmbok7_principle',
    'Methodology': 'methodology',
    'Methodology detail': 'methodology_detail',
    '2026 ECO Domain': 'eco2026_domain',
    '2026 ECO Task': 'eco2026_task',
    'PMBOK8 Performance Domain': 'pmbok8_domain',
    'PMBOK8 Focus Area': 'pmbok8_focus_area',
    'PMBOK8 Principle': 'pmbok8_principle',
    'PMBOK8 Process': 'pmbok8_process',
    'PMBOK8 New Topics': 'pmbok8_new_topics',
    'Question_KR': 'question_kr',
    'A_KR': 'opt_a_kr',
    'B_KR': 'opt_b_kr',
    'C_KR': 'opt_c_kr',
    'D_KR': 'opt_d_kr',
    'E_KR': 'opt_e_kr',
    'Explanation_KR': 'explanation_kr',
}

# Column mapping from translated Excel (PMP_Raw_translated.xlsx) to model fields
# Keys must match the headers written by the translation schedule task.
TRANSLATION_COLUMN_MAP = {
    'No': 'no',
    'Question_ZH': 'question_zh', 'OptA_ZH': 'opt_a_zh', 'OptB_ZH': 'opt_b_zh',
    'OptC_ZH': 'opt_c_zh', 'OptD_ZH': 'opt_d_zh', 'OptE_ZH': 'opt_e_zh',
    'Explanation_ZH': 'explanation_zh',
    'Question_ES': 'question_es', 'OptA_ES': 'opt_a_es', 'OptB_ES': 'opt_b_es',
    'OptC_ES': 'opt_c_es', 'OptD_ES': 'opt_d_es', 'OptE_ES': 'opt_e_es',
    'Explanation_ES': 'explanation_es',
    'Question_JA': 'question_ja', 'OptA_JA': 'opt_a_ja', 'OptB_JA': 'opt_b_ja',
    'OptC_JA': 'opt_c_ja', 'OptD_JA': 'opt_d_ja', 'OptE_JA': 'opt_e_ja',
    'Explanation_JA': 'explanation_ja',
}


def _read_translation_index(filepath):
    """Read PMP_Raw_translated.xlsx into a dict keyed by question No."""
    if not filepath or not os.path.exists(filepath):
        print(f"[translation] No translation file at {filepath} (skipping).")
        return {}
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active  # first sheet (typically 'PMP_Translated')
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_indices = {}
    for excel_col, model_field in TRANSLATION_COLUMN_MAP.items():
        if excel_col in headers:
            col_indices[model_field] = headers.index(excel_col)
    index = {}
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        no_val = values[col_indices.get('no', 0)]
        if not no_val:
            continue
        try:
            no_int = int(no_val)
        except (TypeError, ValueError):
            continue
        entry = {}
        for model_field, col_idx in col_indices.items():
            if model_field == 'no':
                continue
            val = values[col_idx]
            if val is not None:
                entry[model_field] = str(val).strip() if isinstance(val, str) else val
        index[no_int] = entry
    wb.close()
    print(f"[translation] Loaded {len(index)} translated rows from {filepath}.")
    return index

def load_questions(filepath, translation_filepath=None):
    """Load questions from main Excel + optional translation Excel.

    Args:
      filepath: path to PMP_Raw.xlsx (English source + KR columns).
      translation_filepath: optional path to PMP_Raw_translated.xlsx
        (ZH/ES/JA columns). If None or missing, looks for
        'data/PMP_Raw_translated.xlsx' next to the main file.
    """
    print(f"Loading questions from: {filepath}")

    # Resolve translation file path
    if translation_filepath is None:
        guess = os.path.join(os.path.dirname(filepath), 'PMP_Raw_translated.xlsx')
        if os.path.exists(guess):
            translation_filepath = guess
    translation_index = _read_translation_index(translation_filepath) if translation_filepath else {}

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['PMP_All_Data']

    # Get header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find Answer column (first one at index 7)
    answer_col_idx = 7  # First "Answer" column

    # Build column index map
    col_indices = {}
    for excel_col, model_field in COLUMN_MAP.items():
        if excel_col in headers:
            col_indices[model_field] = headers.index(excel_col)

    loaded = 0
    skipped = 0

    with app.app_context():
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]

            # Get question number
            no_val = values[col_indices.get('no', 0)]
            if not no_val:
                skipped += 1
                continue

            # Get answer from first Answer column
            answer_val = values[answer_col_idx]
            if not answer_val:
                skipped += 1
                continue

            # Build question data
            q_data = {}
            for model_field, col_idx in col_indices.items():
                val = values[col_idx]
                if val is not None:
                    q_data[model_field] = str(val).strip() if isinstance(val, str) else val

            q_data['no'] = int(no_val)
            q_data['answer'] = str(answer_val).strip()

            # Merge translation columns if available
            translation = translation_index.get(q_data['no'])
            if translation:
                q_data.update(translation)

            # Check if question exists
            existing = Question.query.filter_by(no=q_data['no']).first()
            if existing:
                for key, val in q_data.items():
                    setattr(existing, key, val)
            else:
                q = Question(**q_data)
                db.session.add(q)

            loaded += 1
            if loaded % 100 == 0:
                print(f"  Loaded {loaded} questions...")
                db.session.commit()

        db.session.commit()

    wb.close()
    print(f"Done! Loaded: {loaded}, Skipped: {skipped}, Translations merged: {len(translation_index)}")
    return loaded

if __name__ == '__main__':
    filepath = sys.argv[1] if len(sys.argv) > 1 else 'data/PMP_Raw.xlsx'
    translation_filepath = sys.argv[2] if len(sys.argv) > 2 else None
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)
    load_questions(filepath, translation_filepath)

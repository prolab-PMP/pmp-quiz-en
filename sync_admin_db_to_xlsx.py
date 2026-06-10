"""Sync admin-edited PMP EN questions from an admin export to PMP_Raw.xlsx.

Usage:
    python sync_admin_db_to_xlsx.py --json admin_questions.json
    python sync_admin_db_to_xlsx.py --json admin_questions.json --dry-run

The script updates the Excel workbook from exported question rows. It creates a
timestamped workbook backup and a CSV change report before saving.
"""
from __future__ import annotations

import argparse
import csv
import json
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SHEET_NAME = "PMP_All_Data"

FIELD_TO_HEADER = {
    "question": "Question. \ubc88\uc5ed, \ucf54\ub4dc,Web\uc73c\ub85c \ub9cc\ub4e4\uae30",
    "opt_a": "A",
    "opt_b": "B",
    "opt_c": "C",
    "opt_d": "D",
    "opt_e": "E",
    "answer": "Answer",
    "explanation": "Explanation(Changed)",
    "eco2021_domain": "2021 ECO Domain",
    "eco2021_task": "2021 ECO Task",
    "pmbok7_domain": "PMBOK7 Performance Domain",
    "pmbok7_principle": "PMBOK7 Principle",
    "methodology": "Methodology",
    "methodology_detail": "Methodology detail",
    "eco2026_domain": "2026 ECO Domain",
    "eco2026_task": "2026 ECO Task",
    "pmbok8_domain": "PMBOK8 Performance Domain",
    "pmbok8_focus_area": "PMBOK8 Focus Area",
    "pmbok8_principle": "PMBOK8 Principle",
    "pmbok8_process": "PMBOK8 Process",
    "pmbok8_new_topics": "PMBOK8 New Topics",
    "question_kr": "Question_KR",
    "opt_a_kr": "A_KR",
    "opt_b_kr": "B_KR",
    "opt_c_kr": "C_KR",
    "opt_d_kr": "D_KR",
    "opt_e_kr": "E_KR",
    "explanation_kr": "Explanation_KR",
}


def norm(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="data/PMP_Raw.xlsx", help="Workbook to update")
    parser.add_argument("--report-dir", default="data", help="Directory for CSV report")
    parser.add_argument("--json", default=None, help="Question JSON exported from the admin site")
    parser.add_argument(
        "--sqlite",
        default=None,
        help="Optional local sqlite DB path. JSON is preferred for production admin edits.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Report only; do not save workbook")
    return parser.parse_args()


def load_questions_from_json(path: str) -> tuple[str, dict[int, dict]]:
    with open(path, "r", encoding="utf-8") as f:
        rows = json.load(f)
    return path, {int(row["no"]): row for row in rows}


def load_questions_from_sqlite(path: str, field_names: list[str]) -> tuple[str, dict[int, dict]]:
    db_path = Path(path).resolve()
    conn = sqlite3.connect(f"file:{db_path}?mode=ro&immutable=1", uri=True)
    conn.row_factory = sqlite3.Row
    try:
        cols = ["no", *field_names]
        rows = conn.execute(f"select {', '.join(cols)} from questions").fetchall()
        return str(db_path), {int(row["no"]): dict(row) for row in rows}
    finally:
        conn.close()


def load_questions(args, field_names: list[str]) -> tuple[str, dict[int, dict]]:
    if args.json:
        return load_questions_from_json(args.json)
    if args.sqlite:
        return load_questions_from_sqlite(args.sqlite, field_names)
    raise SystemExit("Provide --json exported from /admin/questions/export.json")


def main():
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = xlsx_path.with_name(f"{xlsx_path.name}.bak.admin-sync.{timestamp}")
    report_dir = Path(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / f"admin_db_to_xlsx_changes.{timestamp}.csv"

    wb = load_workbook(xlsx_path)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in ws[1]]
    header_to_col = {}
    for idx, header in enumerate(headers, start=1):
        header_to_col.setdefault(header, idx)

    no_col = header_to_col.get("No")
    if not no_col:
        raise SystemExit("No column not found")

    field_to_col = {
        field: header_to_col.get(header)
        for field, header in FIELD_TO_HEADER.items()
        if header_to_col.get(header)
    }
    field_names = sorted(field_to_col)
    source_label, db_questions = load_questions(args, field_names)

    changes = []
    questions_seen = 0

    for row_idx in range(2, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=no_col).value
        if not no:
            continue
        try:
            no = int(no)
        except (TypeError, ValueError):
            continue

        q = db_questions.get(no)
        if not q:
            continue
        questions_seen += 1

        for field, col_idx in field_to_col.items():
            old_value = norm(ws.cell(row=row_idx, column=col_idx).value)
            new_raw = q.get(field)
            new_value = norm(new_raw)
            if old_value != new_value:
                changes.append({
                    "no": no,
                    "field": field,
                    "header": FIELD_TO_HEADER[field],
                    "old": old_value,
                    "new": new_value,
                })
                ws.cell(row=row_idx, column=col_idx).value = new_raw if new_raw is not None else None

    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["no", "field", "header", "old", "new"])
        writer.writeheader()
        writer.writerows(changes)

    if not args.dry_run and changes:
        shutil.copy2(xlsx_path, backup_path)
        wb.save(xlsx_path)
    wb.close()

    print(f"Source: {source_label}")
    print(f"Workbook: {xlsx_path}")
    print(f"Questions matched: {questions_seen}")
    print(f"Fields changed: {len(changes)}")
    print(f"Report: {report_path}")
    if args.dry_run:
        print("Dry run: workbook not saved")
    elif changes:
        print(f"Backup: {backup_path}")
        print("Workbook saved")
    else:
        print("No workbook changes needed")


if __name__ == "__main__":
    main()
